"""
Generate Excel from results.json files in logs/. Output format mirrors combined_results.xlsx.
Output:
  logs/no_case_results.xlsx
  logs/type_help_results.xlsx
"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).parent.parent
LOGS  = _ROOT / "logs"

COLS = ["model", "story", "Overall", "Acc", "Comp.", "Depth",
        "D1(S)", "D2(C)", "D3(X)", "Inst.", "Cit", "Read.", "n_qa"]

# Colors per story group (alternating pair), matching original table
STORY_COLORS = [
    ("FFDDEEFF", "FFCFE0F1"),   # blue
    ("FFE2EFDA", "FFD4E1CC"),   # green
    ("FFFFF2CC", "FFF1E4BE"),   # yellow
    ("FFFCE4D6", "FFFAD0BC"),   # orange
    ("FFF2CEEF", "FFE8B8E4"),   # purple
]

HEADER_FILL  = PatternFill("solid", fgColor="FF1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFFFF", size=11)
OVERALL_FONT_COLOR = "FFC00000"   # red
INST_FONT_COLOR    = "FF999999"   # gray

COL_WIDTHS = {
    "model": 26, "story": 30, "Overall": 9, "Acc": 8,
    "Comp.": 8, "Depth": 8, "D1(S)": 8, "D2(C)": 8, "D3(X)": 8,
    "Inst.": 8, "Cit": 8, "Read.": 8, "n_qa": 7,
}


def load_row(results_path: Path) -> dict:
    r = json.load(open(results_path))

    acc  = r["accuracy"]["pct"]
    comp = r.get("comp_accuracy", {}).get("pct", 0.0)

    depth_map  = r.get("accuracy_by_depth", {})
    d1 = depth_map.get("1_Surface",       {}).get("pct", 0.0)
    d2 = depth_map.get("2_Character",     {}).get("pct", 0.0)
    d3 = depth_map.get("3_Cross-section", {}).get("pct", 0.0)
    # depth composite: average across however many levels exist
    active = [v for v in [d1, d2, d3] if v is not None and
              any(k in depth_map for k in ["1_Surface","2_Character","3_Cross-section"])]
    depth_avg = (d1 + d2 + d3) / 3

    inst  = r.get("inst_stats", {}).get("pass_rate", 0.0)
    cit_d = r.get("cit_score_distribution", {})
    n_qa  = len(r["qa_results"])
    cit   = (cit_d.get("HIGH", 0) * 1.0 + cit_d.get("MEDIUM", 0) * 0.5) / n_qa * 100 if n_qa else 0
    read  = r.get("read_coverage", {}).get("pct", 0.0)

    weights = {"Acc": 3.0, "Read.": 3.0, "Comp.": 1.5, "Depth": 1.5, "Inst.": 0.5, "Cit": 0.5}
    total_w = sum(weights.values())
    overall = (acc * 3.0 + read * 3.0 + comp * 1.5 +
               depth_avg * 1.5 + inst * 0.5 + cit * 0.5) / total_w

    return {
        "model":   r["model"],
        "story":   r["story"],
        "Overall": round(overall, 2),
        "Acc":     round(acc, 2),
        "Comp.":   round(comp, 2),
        "Depth":   round(depth_avg, 2),
        "D1(S)":   round(d1, 2),
        "D2(C)":   round(d2, 2),
        "D3(X)":   round(d3, 2),
        "Inst.":   round(inst, 2),
        "Cit":     round(cit, 2),
        "Read.":   round(read, 2),
        "n_qa":    n_qa,
    }


def build_excel(rows: list[dict], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # header row
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # group by story, assign colors
    story_order = []
    for row in rows:
        if row["story"] not in story_order:
            story_order.append(row["story"])

    story_color_idx = {s: i % len(STORY_COLORS) for i, s in enumerate(story_order)}

    # alternate between two colors within each story group
    story_row_counter: dict[str, int] = {}

    for ri, row in enumerate(rows, 2):
        story = row["story"]
        ci_idx = story_color_idx[story]
        cnt = story_row_counter.get(story, 0)
        story_row_counter[story] = cnt + 1
        color_pair = STORY_COLORS[ci_idx]
        fill_color = color_pair[cnt % 2]
        row_fill = PatternFill("solid", fgColor=fill_color)

        for ci, col in enumerate(COLS, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill

            if col in ("model", "story"):
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(size=10, color="FF000000")
            elif col == "n_qa":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=10, color="FF000000")
            elif col == "Overall":
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=10, color=OVERALL_FONT_COLOR)
            elif col == "Inst.":
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=10, color=INST_FONT_COLOR)
            else:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=10, color="FF000000")

    # column widths
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 9)

    wb.save(out_path)
    print(f"  → {out_path}  ({len(rows)} rows)")


def main():
    for game in ("no_case_should_remain_unsolved", "type_help"):
        game_dir = LOGS / game
        if not game_dir.exists():
            print(f"[skip] {game_dir} not found")
            continue

        rows = []
        for session_dir in sorted(game_dir.iterdir()):
            rp = session_dir / "results.json"
            if rp.exists():
                try:
                    rows.append(load_row(rp))
                except Exception as e:
                    print(f"  [warn] {rp}: {e}")

        if not rows:
            print(f"[skip] {game}: no results.json found")
            continue

        out = LOGS / f"{game}_results.xlsx"
        print(f"{game}: {len(rows)} entries")
        build_excel(rows, out)


if __name__ == "__main__":
    main()
